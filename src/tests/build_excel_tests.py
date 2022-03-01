# DIP Test Excel pop
import openpyxl
import json
import os
import os.path
from os.path import abspath, dirname


# read in the emply template using openpyxl
# get the directory for 'samples' --
SAMPLES_DIR  = os.path.join(dirname(abspath( __file__ )) , "../../docs/samples")
EXCEL_TEMPLATE   = os.path.join(SAMPLES_DIR, "DIP_Excel_Template.xlsx")

workbook = openpyxl.load_workbook(EXCEL_TEMPLATE)
#workbook = openpyxl.load_workbook('DIP_Excel_Template.xlsx')
ws = workbook["Inventory"]
column_list = [cell.value for cell in ws[1]]
#print(column_list)

def findExcelColumn(fieldName):
    for cell in ws[1]:
        if(cell.value == fieldName):
            #print(cell.value)
            return cell.column


# read in a test file (later this will generate from build_test)
with open('test-2.json') as json_file:
    parsed = json.load(json_file)

for o, v in parsed.items():
    if(o.rfind('#')>0):
        start = o.rfind('#')
    else:
        start = o.rfind('/')
    term = o[start+1:]
    if term == 'identifier':
        term = 'unique identifier'
    colNo = findExcelColumn(term)

    # -- check for values that need stripping -- 
    if not term.startswith('@'):
        #print(term)
        #print(findExcelColumn(term))
        #print(v)
        if not isinstance(v, str):
            if isinstance(v, list):
                pvalue = ",".join(v)
                print('its a list!  ' + pvalue)
            else:
                #print('stuck here -- ' + str(v))
                if not v.get("@value"):
                    if term == 'characteristics': 
                        #print('this must be a characteritic! ' + term)
                        for i,j in v.items():
                            if i == '@type':
                                pass
                            else:
                                colNo = findExcelColumn(i)
                                #print('foud col: ' + str(colNo))
                                #print(i)
                                pvalue = j
                                ws.cell(row=2, column=colNo).value = pvalue
                    elif term in ['publisher','creator','governance','owner','steward','custodian','contactPoint']: 
                        #print('What is this? ' + term)
                        for i,j in v.items():
                            if i == '@type':
                                pass
                            else:
                                if ('#fn' in i) | ('#organization-unit' in i):
                                    pvalue = j
                                    #print('found pvalue: ' + pvalue)
                                    #print('foud col: ' + str(colNo))
                                    ws.cell(row=2, column=colNo).value = pvalue
                    else:
                        print('******************* ERROR *********************')
                        print('Could not find @value for attribute: ' + term)                        
                else:
                    pvalue = v["@value"]
                    if (isinstance(pvalue, str)): 
                        if ('<http' in pvalue):
                            pvalue = pvalue.replace('<','').replace('>','')
                        # check for and fix date
                        elif(len(pvalue.split('-')) == 3):
                            #print('herrreeeeeee! ' + str(len(pvalue.split('-'))))
                            pdate = pvalue.split('-')
                            pvalue = pdate[1] + '/' + pdate[2] + '/' + pdate[0]
            ws.cell(row=2, column=colNo).value = pvalue
        else:
            if not colNo:
                print('******************* ERROR *********************')
                print('Could not find column for attribute: ' + term)
                print('Please make sure you are using the correct version of the Excel Templage!')
                break
            ws.cell(row=2, column=colNo).value = v


workbook.save("test_excel-1.xlsx")

