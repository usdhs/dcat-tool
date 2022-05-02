#!/usr/bin/env python3
"""
List datasets and other information

"""

import os
import os.path
from os.path import abspath, dirname
import sys
import pprint
import glob
import time
import json

import openpyxl
import rdflib
from rdflib import Dataset, Graph, URIRef, Literal, Namespace, BNode

import jinja2
import template_reader
import dhs_ontology
import easy_workbook

INSTRUCTIONS  = os.path.join(dirname(abspath( __file__ )) , "instructions.md")

def getCollectVersion():
    gr = dhs_ontology.dhs_collect_graph()
    pred = URIRef("http://www.w3.org/2002/07/owl#versionInfo")
    vers = gr.objects(predicate=pred)
    finalVer = ''
    for foundver in vers:
        finalVer = foundver
    return finalVer


def make_template(v, fname, include_instructions):
    eg = easy_workbook.ExcelGenerator()
    if include_instructions:
        dhsCollectVers = getCollectVersion()
        #print(dhsCollectVers)
        eg.add_markdown_sheet("Instructions", open(INSTRUCTIONS).read(), dhsCollectVers)
    eg.add_columns_sheet("Inventory", v.ci_objs)
    eg.save( fname )

if __name__=="__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Use a DCAT schema to produce capture instruments and APIs.',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--schemata_dir",
                        help="specify the directory to read all the schemata. If the --schema file is in schemata_dir, it will only be read once.",
                        default=dhs_ontology.SCHEMATA_DIR)
    parser.add_argument("--schema_file",
                        help="specify the collection schema file in Turtle, RDF/XML or RDF/JSON",
                        default=dhs_ontology.COLLECT_TTL)
    parser.add_argument("--debug", help="Enable debugging", action='store_true')
    parser.add_argument("--dumpts", help="Dump the triple store after everything it is read", action='store_true')
    parser.add_argument("--dumpci", help="Dump the collection instrument after everything it is read", action='store_true')
    parser.add_argument("--writeschema", help="write the schema to the specified file")
    parser.add_argument("--render_descriptions", help="Render descriptions of each attribute in an HTML file")
    parser.add_argument("--render_excel_descriptions", help="Render descriptions of each attribute in an Excel file")
    parser.add_argument("--render_namespace", help="Render namespace descriptions for novel attributes in an HTML file")
    parser.add_argument("--make_template",
                        help="specify the output filename of the Excel template to make for a collection schema")
    parser.add_argument("--read_xlsx", help="Read a filled-out Excel template and generate multi-line output JSON for each without validating.")
    parser.add_argument("--noinstructions", help="Do not generate instructions. Mostly for debugging.",
                        action='store_true')
    parser.add_argument("--validate_xlsx", help="Validate a filled-out Excel template and generate validateion report")
    parser.add_argument("--validate", help="Read a single JSON object on stdin and validate.", action='store_true')
    parser.add_argument("--validate_lines", help="Read multiple JSON objects on stdin and validate all.",
                        action='store_true')
    parser.add_argument("--flip", help="Flip exit code with --validate_lines", action='store_true')
    parser.add_argument("--convertJSON", help="Read simplified JSON on stdin line-by-line and output full RDF/JSON using DCATv3 spec",action='store_true')
    args = parser.parse_args()

    if args.dumpts:
        print("Dumping triple store:\n")
        for stmt in sorted( dhs_ontology.dcatv3_ontology(args.schemata_dir, args.schema_file) ):
            pprint.pprint(stmt)
            print()

    if args.read_xlsx:
        for r in dhs_ontology.read_xlsx( args.read_xlsx):
            print( json.dumps(r, default=str) )

    # All of those that follow require 'v':
    v = dhs_ontology.Validator(schemata_dir = args.schemata_dir, schema_file=args.schema_file, debug=args.debug)

    if args.dumpci:
        print("Dumping collection instrument:\n")
        for (ct,obj) in enumerate(v.ci_objs):
            print(ct,str(obj).replace("\n"," "))

    if args.validate:
        data = sys.stdin.readline()
        try:
            jin  = json.loads( data )
        except json.decoder.JSONDecodeError as e:
            print(e)
            print("offending input: ",data)
        try:
            v.validate( jin )
            print("OK")
        except dhs_ontology.ValidationFail as e:
            print("FAIL:"+e.message)

    if args.make_template:
        make_template(v, args.make_template, not args.noinstructions)

    if args.render_descriptions:
        desc = v.get_descriptions()
        env = jinja2.Environment(
            loader = jinja2.PackageLoader('dcat_tool','templates')
        )
        template = env.get_template('definitions.html')
        open(args.render_descriptions,"w").write(template.render(desc = desc))

    if args.render_excel_descriptions:
        desc = v.get_descriptions()
        wb = easy_workbook.Workbook()
        ws = wb.active
        ws.title = "DIP Attribute Descriptions"
        #write header row
        bold = easy_workbook.Font(color="FF0000", bold=True)
        ws['A1'] = 'Attribute Name'
        ws['B1'] = 'Qualified Name'
        ws['C1'] = 'Description'
        ws['D1'] = 'Required'
        ws['E1'] = 'Data Type'
        ws['F1'] = 'Namespace'
        ws['G1'] = 'Group'
        #row = ws.row_dimensions[1]
        #row.font = easy_workbook.BOLD
        ws['A1'].font = easy_workbook.BOLD
        ws['A1'].fill = easy_workbook.STATE3_FILL
        ws['B1'].font = easy_workbook.BOLD
        ws['B1'].fill = easy_workbook.STATE3_FILL
        ws['C1'].font = easy_workbook.BOLD
        ws['C1'].fill = easy_workbook.STATE3_FILL
        ws['D1'].font = easy_workbook.BOLD
        ws['D1'].fill = easy_workbook.STATE3_FILL
        ws['E1'].font = easy_workbook.BOLD
        ws['E1'].fill = easy_workbook.STATE3_FILL
        ws['F1'].font = easy_workbook.BOLD
        ws['F1'].fill = easy_workbook.STATE3_FILL
        ws['G1'].font = easy_workbook.BOLD
        ws['G1'].fill = easy_workbook.STATE3_FILL
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 34
        rowCounter = 2
        for d in desc:
            ws.cell(row=rowCounter,column=1).value = d[3]
            ws.cell(row=rowCounter,column=2).value = d[1]
            ws.cell(row=rowCounter,column=3).value = d[2]
            ws.cell(row=rowCounter,column=4).value = d[5]
            if d[5] == 'Yes':
                ws.cell(row=rowCounter,column=4).fill = easy_workbook.PINK_FILL
            ws.cell(row=rowCounter,column=5).value = d[7]
            ws.cell(row=rowCounter,column=6).value = d[4]
            ws.cell(row=rowCounter,column=6).hyperlink = d[4]
            ws.cell(row=rowCounter,column=6).font = easy_workbook.Font(color="0000FF")
            ws.cell(row=rowCounter,column=7).value = d[0]
            ws.cell(row=rowCounter,column=8).value = rowCounter - 1
            rowCounter += 1
            #print(d[0])
        #print('Done!')
        wb.save(args.render_excel_descriptions)


    if args.render_namespace:
        namesp = v.get_namespace()
        namesp2 = v.get_namespace()
        env = jinja2.Environment(
            loader = jinja2.PackageLoader('dcat_tool','templates')
        )
        template = env.get_template('dhsnamespace.html')
        open(args.render_namespace,"w").write(template.render(desc = namesp, desc2 = namesp2))

    if args.validate_lines:
        """Read lines of JSON, turn each one into a dictionary, then validate them all"""
        records = []
        fail = []
        line = 0
        while True:
            data = sys.stdin.readline()
            if len(data) == 0:
                break
            line += 1
            try:
                records.append( json.loads( data ) )
            except json.decoder.JSONDecodeError as e:
                fail.append([line, e.message])
                continue
        ret = dhs_ontology.validate_inventory_records( v, records)
        if ret['response']==200 and fail==[]:
            print("OK")
        else:
            print("FAILURE:")
            print(json.dumps(ret,indent=4, default=str))
            exit(1 if not args.flip else 0)
        exit(0 if not args.flip else 1)

    if args.validate_xlsx:
        print(json.dumps( dhs_ontology.validate_xlsx( v, args.validate_xlsx), indent=4, default=str))

    if args.writeschema:
        fmt = os.path.splitext(args.writeschema)[1][1:].lower()
        if fmt=='json':
            fmt='json-ld'
        v.g2.serialize(destination=args.writeschema, format=fmt)

    """
    This code is not working propertly.

Test:
echo '{ "dct:identifier": "id102", "dct:title": "This is ID102", "dct:description": "This is the third dataset" }' |  python dcat_tool.py --convertJSON

URLs for help:
https://github.com/RDFLib/rdflib/issues/543
https://stackoverflow.com/questions/65818401/namespace-binding-in-rdflib
https://www.programcreek.com/python/example/18799/rdflib.Namespace
https://stackoverflow.com/questions/28503628/force-rdflib-to-define-a-namespace
https://rdflib.readthedocs.io/en/4.0/_modules/rdflib/namespace.html
https://rdflib.readthedocs.io/en/stable/_modules/rdflib/namespace.html
https://github.com/RDFLib/rdflib/issues/1232
    """

    if args.convertJSON:
        g2 = v.cleanGraph()
        while True:
            data = sys.stdin.readline()
            if len(data) == 0:
                break
            obj = json.loads( data )
            bnode = BNode()     # a GUID is generated
            for (k,v) in obj.items():
                g2.add( ( bnode, URIRef(k), Literal(v)) )
            print(g2)
            print(g2.serialize( format='turtle'))
